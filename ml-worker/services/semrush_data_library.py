"""
Semrush Data Library - AI Orchestration Layer
Albeni 1905 - Invisible Luxury Ecosystem

Imports, stores, and indexes Semrush CSV/XLSX exports to power
the SEO agents with historical and research data.

Supported Semrush export types:
- Keyword Overview / Keyword Magic Tool
- Domain Analytics (Organic/Paid)
- Organic Research (Positions, Competitors, Pages)
- Advertising Research
- Keyword Gap
- Backlink Analytics
- Position Tracking
- Site Audit
- Custom exports
"""

import csv
import io
import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

# Directory for stored Semrush data
SEMRUSH_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "semrush")


class SemrushDataLibrary:
    """
    Parses, indexes and serves Semrush export data.
    Data is stored as JSON for fast querying by agents.
    """

    # Auto-detect Semrush export type by column headers
    EXPORT_SIGNATURES = {
        "keyword_overview": ["Keyword", "Volume", "Keyword Difficulty"],
        "keyword_magic": ["Keyword", "Volume", "KD%"],
        "organic_research_positions": ["Keyword", "Position", "Search Volume", "URL"],
        "organic_research_competitors": ["Domain", "Competition Level", "Common Keywords"],
        "organic_research_pages": ["URL", "Traffic", "Traffic (%)"],
        "advertising_research": ["Keyword", "Position", "CPC", "Title", "Description"],
        "keyword_gap": ["Keyword", "Volume", "KD"],
        "backlink_overview": ["Referring Domain", "Authority Score"],
        "backlink_referring_domains": ["Domain", "Domain Authority Score"],
        "position_tracking": ["Keyword", "Position", "Previous Position"],
        "site_audit": ["Issue", "Pages", "Severity"],
        "domain_overview": ["Domain", "Organic Keywords", "Organic Traffic"],
    }

    # Column name normalization map (Semrush uses different names across reports)
    COLUMN_NORMALIZE = {
        "search volume": "volume",
        "volume": "volume",
        "nq": "volume",
        "keyword difficulty": "kd",
        "keyword difficulty %": "kd",
        "kd%": "kd",
        "kd": "kd",
        "cpc": "cpc",
        "cp": "cpc",
        "cpc (usd)": "cpc",
        "competition": "competition",
        "co": "competition",
        "com.": "competition",
        "position": "position",
        "po": "position",
        "pos.": "position",
        "keyword": "keyword",
        "ph": "keyword",
        "url": "url",
        "traffic": "traffic",
        "traffic (%)": "traffic_pct",
        "tr": "traffic_pct",
        "traffic cost": "traffic_cost",
        "tc": "traffic_cost",
        "domain": "domain",
        "dn": "domain",
        "number of results": "results",
        "nr": "results",
        "results": "results",
        "trend": "trend",
        "td": "trend",
        "serp features": "serp_features",
        "sf": "serp_features",
        "intent": "intent",
        "authority score": "authority_score",
        "ascore": "authority_score",
        "title": "title",
        "tt": "title",
        "description": "description",
        "ds": "description",
        "common keywords": "common_keywords",
        "np": "common_keywords",
        "competition level": "competition_level",
        "cr": "competition_level",
        "previous position": "prev_position",
    }

    # Storage: Postgres (tabella semrush_library_imports). Fino al 2026-07-15 era
    # file-based in SEMRUSH_DATA_DIR: il filesystem del container Railway è
    # effimero, quindi la libreria si svuotava a ogni redeploy. La directory
    # resta solo per l'auto-migrazione di eventuali file legacy.
    _table_ready = False

    def __init__(self):
        self._index = self._load_index()

    def _db(self):
        from models.database import SessionLocal
        return SessionLocal()

    def _ensure_table(self, db) -> None:
        if SemrushDataLibrary._table_ready:
            return
        from sqlalchemy import text
        db.execute(text(
            "CREATE TABLE IF NOT EXISTS semrush_library_imports ("
            "id TEXT PRIMARY KEY, "
            "record JSONB NOT NULL, "
            "data JSONB NOT NULL, "
            "imported_at TIMESTAMPTZ DEFAULT NOW())"
        ))
        db.commit()
        SemrushDataLibrary._table_ready = True

    def _load_index(self) -> List[Dict]:
        """Load the data library index from Postgres (+ auto-migrate legacy files)."""
        from sqlalchemy import text
        try:
            db = self._db()
            try:
                self._ensure_table(db)
                rows = db.execute(text(
                    "SELECT record FROM semrush_library_imports ORDER BY imported_at"
                )).fetchall()
                index = [r[0] for r in rows]
                if not index:
                    migrated = self._migrate_legacy_files(db)
                    if migrated:
                        index = migrated
                return index
            finally:
                db.close()
        except Exception as e:
            logger.error(f"Semrush library: load index from DB failed: {e}")
            return []

    def _migrate_legacy_files(self, db) -> List[Dict]:
        """One-shot: importa in Postgres eventuali file della vecchia libreria file-based."""
        from sqlalchemy import text
        legacy_index = os.path.join(SEMRUSH_DATA_DIR, "_index.json")
        if not os.path.exists(legacy_index):
            return []
        try:
            with open(legacy_index, "r") as f:
                old = json.load(f)
        except Exception:
            return []
        migrated = []
        for rec in old:
            data_path = rec.get("data_path", "")
            if not data_path or not os.path.exists(data_path):
                continue
            try:
                with open(data_path, "r") as f:
                    payload = json.load(f)
                rec = {k: v for k, v in rec.items() if k != "data_path"}
                db.execute(text(
                    "INSERT INTO semrush_library_imports (id, record, data, imported_at) "
                    "VALUES (:id, CAST(:rec AS jsonb), CAST(:data AS jsonb), :ts) "
                    "ON CONFLICT (id) DO NOTHING"
                ), {"id": rec["id"], "rec": json.dumps(rec, default=str),
                    "data": json.dumps(payload.get("data", []), default=str),
                    "ts": rec.get("imported_at")})
                migrated.append(rec)
            except Exception as e:
                logger.warning(f"Semrush library: legacy migration failed for {rec.get('id')}: {e}")
        if migrated:
            db.commit()
            logger.info(f"Semrush library: migrated {len(migrated)} legacy file imports to Postgres")
        return migrated

    # ================================================================
    # IMPORT / PARSE
    # ================================================================

    def parse_csv(self, content: str, filename: str = "") -> Tuple[str, List[Dict]]:
        """
        Parse a Semrush CSV export. Auto-detects the report type.
        Returns (export_type, parsed_rows).
        """
        # Semrush CSVs can use comma or semicolon as delimiter
        delimiter = ";"
        if content.count(",") > content.count(";"):
            delimiter = ","

        # Some Semrush exports have metadata rows at the top — skip them
        lines = content.strip().split("\n")
        header_row_idx = 0
        for i, line in enumerate(lines):
            # Header row typically has many columns and known column names
            cols = line.split(delimiter)
            if len(cols) >= 3 and any(
                c.strip().strip('"').lower() in self.COLUMN_NORMALIZE
                for c in cols[:5]
            ):
                header_row_idx = i
                break

        clean_content = "\n".join(lines[header_row_idx:])
        reader = csv.DictReader(io.StringIO(clean_content), delimiter=delimiter)

        rows = []
        raw_headers = reader.fieldnames or []
        for row in reader:
            normalized = {}
            for key, value in row.items():
                if key is None:
                    continue
                norm_key = self.COLUMN_NORMALIZE.get(key.strip().strip('"').lower(), key.strip().strip('"').lower().replace(" ", "_"))
                # Try to convert numeric values
                value = value.strip().strip('"') if value else ""
                if norm_key in ("volume", "kd", "position", "results", "traffic", "common_keywords", "authority_score", "prev_position"):
                    try:
                        normalized[norm_key] = int(value.replace(",", "").replace(".", "")) if value and value != "-" else 0
                    except ValueError:
                        normalized[norm_key] = value
                elif norm_key in ("cpc", "competition", "traffic_pct", "traffic_cost", "competition_level"):
                    try:
                        normalized[norm_key] = float(value.replace(",", ".")) if value and value != "-" else 0.0
                    except ValueError:
                        normalized[norm_key] = value
                else:
                    normalized[norm_key] = value
            if normalized:
                rows.append(normalized)

        # Auto-detect export type
        export_type = self._detect_type(raw_headers)
        logger.info(f"Parsed {len(rows)} rows from '{filename}' — detected type: {export_type}")

        return export_type, rows

    def parse_xlsx(self, file_bytes: bytes, filename: str = "") -> Tuple[str, List[Dict]]:
        """
        Parse a Semrush XLSX export using openpyxl.
        Returns (export_type, parsed_rows).
        """
        try:
            import openpyxl
        except ImportError:
            # Fallback: try to install
            import subprocess
            subprocess.check_call(["pip", "install", "openpyxl", "--break-system-packages", "-q"])
            import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        # Find header row
        headers = None
        for row in rows_iter:
            str_row = [str(c).strip() if c else "" for c in row]
            if any(s.lower() in self.COLUMN_NORMALIZE for s in str_row[:5] if s):
                headers = str_row
                break

        if not headers:
            # Use first row as headers
            wb.close()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(next(rows_iter))]

        rows = []
        for row in rows_iter:
            normalized = {}
            for i, value in enumerate(row):
                if i >= len(headers):
                    break
                key = headers[i]
                norm_key = self.COLUMN_NORMALIZE.get(key.lower(), key.lower().replace(" ", "_"))
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    normalized[norm_key] = value
                    continue
                else:
                    value = str(value).strip()

                if norm_key in ("volume", "kd", "position", "results", "traffic", "common_keywords", "authority_score"):
                    try:
                        normalized[norm_key] = int(float(value.replace(",", ""))) if value and value != "-" else 0
                    except ValueError:
                        normalized[norm_key] = value
                elif norm_key in ("cpc", "competition", "traffic_pct", "traffic_cost", "competition_level"):
                    try:
                        normalized[norm_key] = float(value.replace(",", ".")) if value and value != "-" else 0.0
                    except ValueError:
                        normalized[norm_key] = value
                else:
                    normalized[norm_key] = value

            if normalized:
                rows.append(normalized)

        wb.close()
        export_type = self._detect_type(headers)
        logger.info(f"Parsed {len(rows)} rows from XLSX '{filename}' — detected type: {export_type}")
        return export_type, rows

    def _detect_type(self, headers: List[str]) -> str:
        """Auto-detect Semrush export type from column headers."""
        header_lower = [h.strip().strip('"').lower() for h in headers if h]

        best_match = "custom"
        best_score = 0

        for export_type, signature_cols in self.EXPORT_SIGNATURES.items():
            score = sum(1 for sig in signature_cols if sig.lower() in header_lower)
            if score > best_score:
                best_score = score
                best_match = export_type

        return best_match if best_score >= 2 else "custom"

    # ================================================================
    # STORAGE
    # ================================================================

    def store_import(
        self,
        export_type: str,
        rows: List[Dict],
        filename: str,
        metadata: Optional[Dict] = None,
    ) -> Dict:
        """
        Store parsed data and add to index.
        Returns the import record.
        """
        import_id = f"sr_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{len(self._index)}"

        record = {
            "id": import_id,
            "filename": filename,
            "export_type": export_type,
            "rows_count": len(rows),
            "imported_at": datetime.utcnow().isoformat(),
            "metadata": metadata or {},
            "columns": list(rows[0].keys()) if rows else [],
        }

        # Extract summary stats
        if rows:
            kw_count = sum(1 for r in rows if "keyword" in r)
            domain_count = len(set(r.get("domain", "") for r in rows if r.get("domain")))
            avg_volume = 0
            volumes = [r["volume"] for r in rows if isinstance(r.get("volume"), (int, float))]
            if volumes:
                avg_volume = round(sum(volumes) / len(volumes))
            record["summary"] = {
                "keywords": kw_count,
                "domains": domain_count,
                "avg_volume": avg_volume,
                "top_volume_keywords": sorted(
                    [r for r in rows if isinstance(r.get("volume"), (int, float))],
                    key=lambda x: x.get("volume", 0),
                    reverse=True
                )[:5],
            }

        # Save data (Postgres)
        from sqlalchemy import text
        db = self._db()
        try:
            self._ensure_table(db)
            db.execute(text(
                "INSERT INTO semrush_library_imports (id, record, data) "
                "VALUES (:id, CAST(:rec AS jsonb), CAST(:data AS jsonb))"
            ), {"id": import_id, "rec": json.dumps(record, default=str),
                "data": json.dumps(rows, default=str)})
            db.commit()
        finally:
            db.close()

        # Update in-memory index
        self._index.append(record)

        logger.info(f"Stored Semrush import: {import_id} ({len(rows)} rows, type: {export_type})")
        return record

    # ================================================================
    # QUERY / RETRIEVE
    # ================================================================

    def list_imports(self) -> List[Dict]:
        """List all imported Semrush datasets."""
        return [
            {k: v for k, v in rec.items() if k != "data_path"}
            for rec in self._index
        ]

    def get_import(self, import_id: str) -> Optional[Dict]:
        """Get a specific import by ID, including full data."""
        from sqlalchemy import text
        db = self._db()
        try:
            self._ensure_table(db)
            row = db.execute(text(
                "SELECT record, data FROM semrush_library_imports WHERE id = :id"
            ), {"id": import_id}).fetchone()
            if not row:
                return None
            return {"record": row[0], "data": row[1]}
        finally:
            db.close()

    def get_import_data(self, import_id: str) -> List[Dict]:
        """Get just the data rows for an import."""
        full = self.get_import(import_id)
        return full["data"] if full else []

    def delete_import(self, import_id: str) -> bool:
        """Delete an import and its data."""
        from sqlalchemy import text
        db = self._db()
        try:
            self._ensure_table(db)
            res = db.execute(text(
                "DELETE FROM semrush_library_imports WHERE id = :id"
            ), {"id": import_id})
            db.commit()
            deleted = bool(res.rowcount)
        finally:
            db.close()
        if deleted:
            self._index = [r for r in self._index if r["id"] != import_id]
        return deleted

    def search_keywords(self, query: str, import_id: Optional[str] = None) -> List[Dict]:
        """
        Search for keywords across all imports (or a specific one).
        Returns matching rows sorted by volume descending.
        """
        results = []
        query_lower = query.lower()

        targets = self._index if not import_id else [r for r in self._index if r["id"] == import_id]

        for rec in targets:
            data = self.get_import_data(rec["id"])
            for row in data:
                kw = str(row.get("keyword", "")).lower()
                if query_lower in kw:
                    row["_source_import"] = rec["id"]
                    row["_source_file"] = rec["filename"]
                    row["_export_type"] = rec["export_type"]
                    results.append(row)

        # Sort by volume descending
        results.sort(key=lambda x: x.get("volume", 0) if isinstance(x.get("volume"), (int, float)) else 0, reverse=True)
        return results[:100]

    def get_all_keywords(self, min_volume: int = 0, max_kd: int = 100) -> List[Dict]:
        """
        Get all keywords across all imports, with optional filters.
        Useful for the SEO agents to build a unified keyword universe.
        """
        all_kws = {}

        for rec in self._index:
            data = self.get_import_data(rec["id"])
            for row in data:
                kw = row.get("keyword", "")
                if not kw:
                    continue
                vol = row.get("volume", 0) if isinstance(row.get("volume"), (int, float)) else 0
                kd = row.get("kd", 0) if isinstance(row.get("kd"), (int, float)) else 0

                if vol >= min_volume and kd <= max_kd:
                    # Keep the row with highest volume if duplicate
                    if kw not in all_kws or vol > all_kws[kw].get("volume", 0):
                        row["_source"] = rec["filename"]
                        all_kws[kw] = row

        result = sorted(all_kws.values(), key=lambda x: x.get("volume", 0), reverse=True)
        return result

    def get_aggregate_stats(self) -> Dict:
        """
        Aggregate stats across all imported Semrush data.
        """
        total_rows = 0
        total_keywords = set()
        total_domains = set()
        types_count = {}
        all_volumes = []

        for rec in self._index:
            total_rows += rec["rows_count"]
            types_count[rec["export_type"]] = types_count.get(rec["export_type"], 0) + 1

            data = self.get_import_data(rec["id"])
            for row in data:
                if row.get("keyword"):
                    total_keywords.add(row["keyword"].lower())
                if row.get("domain"):
                    total_domains.add(row["domain"].lower())
                vol = row.get("volume", 0)
                if isinstance(vol, (int, float)) and vol > 0:
                    all_volumes.append(vol)

        return {
            "total_imports": len(self._index),
            "total_rows": total_rows,
            "unique_keywords": len(total_keywords),
            "unique_domains": len(total_domains),
            "import_types": types_count,
            "avg_volume": round(sum(all_volumes) / max(1, len(all_volumes))) if all_volumes else 0,
            "max_volume": max(all_volumes) if all_volumes else 0,
            "total_volume_tracked": sum(all_volumes),
        }
